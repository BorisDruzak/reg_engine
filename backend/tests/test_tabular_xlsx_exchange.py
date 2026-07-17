import json
from contextlib import nullcontext
from dataclasses import replace
from datetime import date
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook
from starlette.responses import Response

from app.api.v1.endpoints import import_export as import_export_endpoint
from app.schemas.import_export import TabularCardWorkbookRequest
from app.services import import_export
from app.services.cards import InvalidFieldValueError


def test_tabular_xlsx_v2_template_includes_title_and_creation_metadata() -> None:
    field = SimpleNamespace(id=uuid4(), label="Комментарий", field_type="text")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=uuid4(),
        template=SimpleNamespace(id=uuid4(), name="Заявка"),
        fields=(
            import_export.TabularWorkbookField(
                field=field,
                block=SimpleNamespace(id=uuid4(), title="Основное"),
                header="Комментарий",
            ),
        ),
        organizations=(SimpleNamespace(id=uuid4(), name="Администрация", code="admin"),),
        include_organization_column=True,
        fixed_organization_id=None,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
        import_mode="enrich_global_references",
        work_experience_as_of_date=date(2026, 7, 17),
        title_header="Наименование заявки",
    )

    content = import_export.TabularCardExchangeService(MagicMock())._build_workbook(
        actor_user_id=uuid4(),
        configuration=configuration,
        cards=None,
    )

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook["Карточки"]
    metadata = json.loads(workbook["_registry_engine"]["B1"].value)

    assert [cell.value for cell in sheet[1][:4]] == [
        "№ п/п",
        "Наименование заявки",
        "Организация",
        "Комментарий",
    ]
    assert sheet["B2"].number_format == "@"
    assert metadata["format_version"] == "tabular_card_xlsx_v2"
    assert metadata["import_mode"] == "enrich_global_references"
    assert metadata["work_experience_as_of_date"] == "2026-07-17"
    assert metadata["title_header"] == "Наименование заявки"
    assert metadata["title_required"] is True


def test_tabular_xlsx_export_defaults_v2_as_of_date_when_omitted(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    actor_user_id = uuid4()
    registry_id = uuid4()
    organization_id = uuid4()
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Заявка"),
        fields=(
            import_export.TabularWorkbookField(
                field=SimpleNamespace(id=uuid4(), label="Комментарий", field_type="text"),
                block=SimpleNamespace(id=uuid4(), title="Основное"),
                header="Комментарий",
            ),
        ),
        organizations=(SimpleNamespace(id=organization_id, name="Администрация", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
    )

    class ExportCardService:
        def __init__(self, _session: object) -> None:
            pass

        def list_visible_cards(self, **_kwargs: object) -> list[object]:
            return []

    class AuditService:
        def __init__(self, _session: object) -> None:
            pass

        def record_user_event(self, **_kwargs: object) -> None:
            pass

    service = import_export.TabularCardExchangeService(MagicMock())
    monkeypatch.setattr(import_export, "CardService", ExportCardService)
    monkeypatch.setattr(import_export, "AuditService", AuditService)
    monkeypatch.setattr(
        service,
        "_configuration_for_actor",
        lambda **kwargs: replace(
            configuration,
            work_experience_as_of_date=kwargs["work_experience_as_of_date"],
        ),
    )

    content = service.export_xlsx_for_actor(
        actor_user_id=actor_user_id,
        registry_id=registry_id,
        card_template_id=configuration.template.id,
        field_ids=[configuration.fields[0].field.id],
        organization_ids=[organization_id],
    )

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    metadata = json.loads(workbook["_registry_engine"]["B1"].value)

    assert metadata["work_experience_as_of_date"] == date.today().isoformat()
    assert metadata["importable"] is False


def test_tabular_xlsx_uses_stable_dynamic_columns_when_headers_collide_with_fixed_headers() -> None:
    title_field = SimpleNamespace(
        id=uuid4(),
        label="Название карточки",
        field_type="select",
    )
    organization_field = SimpleNamespace(
        id=uuid4(),
        label="Организация",
        field_type="select",
    )
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=uuid4(),
        template=SimpleNamespace(id=uuid4(), name="Заявка"),
        fields=(
            import_export.TabularWorkbookField(
                field=title_field,
                block=SimpleNamespace(id=uuid4(), title="Основное"),
                header="Название карточки",
            ),
            import_export.TabularWorkbookField(
                field=organization_field,
                block=SimpleNamespace(id=uuid4(), title="Основное"),
                header="Организация",
            ),
        ),
        organizations=(SimpleNamespace(id=uuid4(), name="Администрация", code="admin"),),
        include_organization_column=True,
        fixed_organization_id=None,
        organization_labels={},
        reference_labels={
            title_field.id: {"Первое": uuid4()},
            organization_field.id: {"Второе": uuid4()},
        },
        unit_organization_ids={},
    )

    content = import_export.TabularCardExchangeService(MagicMock())._build_workbook(
        actor_user_id=uuid4(),
        configuration=configuration,
        cards=None,
    )

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook["Карточки"]
    validations = {
        validation.sqref: validation for validation in sheet.data_validations.dataValidation
    }

    assert validations["D2:D101"].formula1 == f"=field_{title_field.id.hex}_choices"
    assert validations["E2:E101"].formula1 == f"=field_{organization_field.id.hex}_choices"


def test_tabular_xlsx_request_defaults_to_strict_and_accepts_enrichment_metadata() -> None:
    payload = {
        "card_template_id": str(uuid4()),
        "field_ids": [str(uuid4())],
        "organization_ids": [str(uuid4())],
    }

    strict = TabularCardWorkbookRequest.model_validate(payload)
    enrich = TabularCardWorkbookRequest.model_validate(
        {
            **payload,
            "import_mode": "enrich_global_references",
            "work_experience_as_of_date": "2026-07-17",
        }
    )

    assert strict.import_mode == "strict"
    assert strict.work_experience_as_of_date is None
    assert enrich.import_mode == "enrich_global_references"
    assert enrich.work_experience_as_of_date == date(2026, 7, 17)


def test_tabular_xlsx_enrichment_plans_once_then_creates_one_global_reference_for_repeated_labels(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    actor_user_id = uuid4()
    registry_id = uuid4()
    organization_id = uuid4()
    reference_list_id = uuid4()
    field = SimpleNamespace(
        id=uuid4(),
        label="Status",
        field_type="select",
        options_source_type="reference_list",
        options_source_id=reference_list_id,
    )
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Cards"),
        fields=(
            import_export.TabularWorkbookField(
                field=field,
                block=SimpleNamespace(id=uuid4(), title="Main"),
                header="Status",
            ),
        ),
        organizations=(SimpleNamespace(id=organization_id, name="Administration", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={field.id: {}},
        unit_organization_ids={},
        import_mode="enrich_global_references",
        work_experience_as_of_date=date(2026, 7, 17),
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки", "Status"])
    sheet.append([1, "First", "  NEW\u00a0 STATUS "])
    sheet.append([2, "Second", "new   status"])
    workbook.create_sheet("_registry_engine")["B1"] = "{}"
    content = BytesIO()
    workbook.save(content)
    created_items: list[object] = []
    written_values: list[object] = []

    class ImportSession:
        def begin_nested(self) -> object:
            return nullcontext()

    class References:
        def __init__(self, _session: object) -> None:
            pass

        def resolve_or_plan_global_import_item_for_actor(self, **kwargs: object) -> object:
            return SimpleNamespace(
                status="create",
                normalized_label="new status",
                display_label="NEW STATUS",
                reference_item_id=None,
            )

        def create_global_import_item_for_actor(self, **kwargs: object) -> object:
            item = SimpleNamespace(id=uuid4(), label=kwargs["display_label"])
            created_items.append(item)
            return item

    class ImportCards:
        def __init__(self, _session: object) -> None:
            pass

        def validate_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

        def create_card_for_actor(self, **_kwargs: object) -> object:
            return SimpleNamespace(id=uuid4())

        def set_field_value_for_actor(self, *, value: object, **_kwargs: object) -> None:
            written_values.append(value)

    class Audit:
        def __init__(self, _session: object) -> None:
            pass

        def record_user_event(self, **_kwargs: object) -> None:
            pass

    service = import_export.TabularCardExchangeService(ImportSession())
    monkeypatch.setattr(import_export, "ReferenceListService", References, raising=False)
    monkeypatch.setattr(import_export, "CardService", ImportCards)
    monkeypatch.setattr(import_export, "AuditService", Audit)
    monkeypatch.setattr(service, "_configuration_from_metadata", lambda **_kwargs: configuration)

    preview = service.preview_import_xlsx_for_actor(
        actor_user_id=actor_user_id,
        registry_id=registry_id,
        xlsx_content=content.getvalue(),
    )
    committed = service.commit_import_xlsx_for_actor(
        actor_user_id=actor_user_id,
        registry_id=registry_id,
        xlsx_content=content.getvalue(),
    )

    assert preview["summary"]["would_create_reference_items"] == 1
    assert preview["new_reference_items"] == [{"field_label": "Status", "label": "NEW STATUS"}]
    assert committed["summary"]["created_reference_items"] == 1
    assert len(created_items) == 1
    assert written_values == [created_items[0].id, created_items[0].id]


def test_tabular_xlsx_strict_unknown_reference_is_invalid_and_never_planned(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry_id = uuid4()
    organization_id = uuid4()
    field = SimpleNamespace(id=uuid4(), field_type="select", label="Status")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Cards"),
        fields=(
            import_export.TabularWorkbookField(
                field=field, block=SimpleNamespace(id=uuid4()), header="Status"
            ),
        ),
        organizations=(SimpleNamespace(id=organization_id, name="Administration", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={field.id: {}},
        unit_organization_ids={},
        import_mode="strict",
        work_experience_as_of_date=date(2026, 7, 17),
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки", "Status"])
    sheet.append([1, "Strict", "unknown"])
    workbook.create_sheet("_registry_engine")["B1"] = "{}"
    content = BytesIO()
    workbook.save(content)

    class ImportCards:
        def __init__(self, _session: object) -> None:
            pass

        def validate_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

    service = import_export.TabularCardExchangeService(MagicMock())
    monkeypatch.setattr(import_export, "CardService", ImportCards)
    monkeypatch.setattr(service, "_configuration_from_metadata", lambda **_kwargs: configuration)

    preview = service.preview_import_xlsx_for_actor(
        actor_user_id=uuid4(), registry_id=registry_id, xlsx_content=content.getvalue()
    )

    assert preview["rows"][0]["status"] == "invalid"
    assert preview["summary"]["would_create_reference_items"] == 0
    assert preview["new_reference_items"] == []


def test_tabular_xlsx_enrichment_rejects_select_without_reference_list_configuration(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry_id = uuid4()
    organization_id = uuid4()
    field = SimpleNamespace(id=uuid4(), field_type="select", label="Status")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Cards"),
        fields=(
            import_export.TabularWorkbookField(
                field=field, block=SimpleNamespace(id=uuid4()), header="Status"
            ),
        ),
        organizations=(SimpleNamespace(id=organization_id, name="Administration", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={field.id: {}},
        unit_organization_ids={},
        import_mode="enrich_global_references",
        work_experience_as_of_date=date(2026, 7, 17),
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки", "Status"])
    sheet.append([1, "No list", "unknown"])
    workbook.create_sheet("_registry_engine")["B1"] = "{}"
    content = BytesIO()
    workbook.save(content)

    class ImportCards:
        def __init__(self, _session: object) -> None:
            pass

        def validate_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

    service = import_export.TabularCardExchangeService(MagicMock())
    monkeypatch.setattr(import_export, "CardService", ImportCards)
    monkeypatch.setattr(service, "_configuration_from_metadata", lambda **_kwargs: configuration)

    preview = service.preview_import_xlsx_for_actor(
        actor_user_id=uuid4(), registry_id=registry_id, xlsx_content=content.getvalue()
    )

    assert preview["rows"][0]["status"] == "invalid"
    assert preview["summary"]["would_create_reference_items"] == 0


@pytest.mark.parametrize(
    "options_config_json",
    [
        {"reference_resolution": "by_card_organization"},
        {"allow_owner_override": True},
    ],
)
def test_tabular_xlsx_enrichment_rejects_organization_aware_reference_resolution(
    monkeypatch: pytest.MonkeyPatch,
    options_config_json: dict[str, object],
) -> None:
    registry_id = uuid4()
    organization_id = uuid4()
    field = SimpleNamespace(
        id=uuid4(),
        field_type="select",
        label="Status",
        options_source_type="reference_list",
        options_source_id=uuid4(),
        options_config_json=options_config_json,
    )
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Cards"),
        fields=(
            import_export.TabularWorkbookField(
                field=field, block=SimpleNamespace(id=uuid4()), header="Status"
            ),
        ),
        organizations=(SimpleNamespace(id=organization_id, name="Administration", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={field.id: {}},
        unit_organization_ids={},
        import_mode="enrich_global_references",
        work_experience_as_of_date=date(2026, 7, 17),
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки", "Status"])
    sheet.append([1, "Organization aware", "unknown"])
    workbook.create_sheet("_registry_engine")["B1"] = "{}"
    content = BytesIO()
    workbook.save(content)

    class References:
        def __init__(self, _session: object) -> None:
            pass

        def resolve_or_plan_global_import_item_for_actor(self, **_kwargs: object) -> object:
            return SimpleNamespace(
                status="create",
                normalized_label="unknown",
                display_label="unknown",
                reference_item_id=None,
            )

    class ImportCards:
        def __init__(self, _session: object) -> None:
            pass

        def validate_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

    service = import_export.TabularCardExchangeService(MagicMock())
    monkeypatch.setattr(import_export, "ReferenceListService", References)
    monkeypatch.setattr(import_export, "CardService", ImportCards)
    monkeypatch.setattr(service, "_configuration_from_metadata", lambda **_kwargs: configuration)

    preview = service.preview_import_xlsx_for_actor(
        actor_user_id=uuid4(), registry_id=registry_id, xlsx_content=content.getvalue()
    )

    assert preview["rows"][0]["status"] == "invalid"
    assert preview["summary"]["would_create_reference_items"] == 0
    assert preview["new_reference_items"] == []


def test_tabular_xlsx_enrichment_rolls_back_created_references_and_cards_after_write_failure(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    actor_user_id = uuid4()
    registry_id = uuid4()
    organization_id = uuid4()
    reference_list_id = uuid4()
    field = SimpleNamespace(
        id=uuid4(),
        field_type="select",
        label="Status",
        options_source_type="reference_list",
        options_source_id=reference_list_id,
        options_config_json={},
    )
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Cards"),
        fields=(
            import_export.TabularWorkbookField(
                field=field, block=SimpleNamespace(id=uuid4()), header="Status"
            ),
        ),
        organizations=(SimpleNamespace(id=organization_id, name="Administration", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={field.id: {}},
        unit_organization_ids={},
        import_mode="enrich_global_references",
        work_experience_as_of_date=date(2026, 7, 17),
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки", "Status"])
    sheet.append([1, "First", "new status"])
    sheet.append([2, "Second", "new status"])
    workbook.create_sheet("_registry_engine")["B1"] = "{}"
    content = BytesIO()
    workbook.save(content)
    persisted_references: list[object] = []
    persisted_cards: list[object] = []
    transactions: list[object] = []

    class Transaction:
        def __enter__(self) -> object:
            return self

        def __exit__(self, exc_type: object, _exc: object, _traceback: object) -> bool:
            if exc_type is not None:
                persisted_references.clear()
                persisted_cards.clear()
            return False

    class ImportSession:
        def begin_nested(self) -> object:
            transaction = Transaction()
            transactions.append(transaction)
            return transaction

    class References:
        def __init__(self, _session: object) -> None:
            pass

        def resolve_or_plan_global_import_item_for_actor(self, **_kwargs: object) -> object:
            return SimpleNamespace(
                status="create",
                normalized_label="new status",
                display_label="new status",
                reference_item_id=None,
            )

        def create_global_import_item_for_actor(self, **_kwargs: object) -> object:
            item = SimpleNamespace(id=uuid4())
            persisted_references.append(item)
            return item

    class ImportCards:
        def __init__(self, _session: object) -> None:
            pass

        def validate_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

        def create_card_for_actor(self, **_kwargs: object) -> object:
            if persisted_cards:
                raise RuntimeError("simulated card write failure")
            card = SimpleNamespace(id=uuid4())
            persisted_cards.append(card)
            return card

        def set_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

    class Audit:
        def __init__(self, _session: object) -> None:
            pass

        def record_user_event(self, **_kwargs: object) -> None:
            pass

    service = import_export.TabularCardExchangeService(ImportSession())
    monkeypatch.setattr(import_export, "ReferenceListService", References)
    monkeypatch.setattr(import_export, "CardService", ImportCards)
    monkeypatch.setattr(import_export, "AuditService", Audit)
    monkeypatch.setattr(service, "_configuration_from_metadata", lambda **_kwargs: configuration)

    with pytest.raises(RuntimeError, match="write failure"):
        service.commit_import_xlsx_for_actor(
            actor_user_id=actor_user_id,
            registry_id=registry_id,
            xlsx_content=content.getvalue(),
        )

    assert len(transactions) == 1
    assert persisted_references == []
    assert persisted_cards == []


def test_tabular_xlsx_rejects_legacy_v1_metadata() -> None:
    with pytest.raises(import_export.ImportExportServiceError, match="Версия XLSX-шаблона"):
        import_export.TabularCardExchangeService(MagicMock())._configuration_from_metadata(
            actor_user_id=uuid4(),
            registry_id=uuid4(),
            metadata={"format_version": "tabular_card_xlsx_v1"},
        )


def test_tabular_xlsx_rejects_export_workbook_as_explicitly_non_importable() -> None:
    with pytest.raises(import_export.ImportExportServiceError, match="не предназначен для импорта"):
        import_export.TabularCardExchangeService(MagicMock())._configuration_from_metadata(
            actor_user_id=uuid4(),
            registry_id=uuid4(),
            metadata={
                "format_version": "tabular_card_xlsx_v2",
                "importable": False,
            },
        )


def test_tabular_xlsx_import_marks_blank_card_title_invalid_and_uses_title_on_commit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    actor_user_id = uuid4()
    registry_id = uuid4()
    organization_id = uuid4()
    field = SimpleNamespace(id=uuid4(), label="Комментарий", field_type="text")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Заявка"),
        fields=(
            import_export.TabularWorkbookField(
                field=field,
                block=SimpleNamespace(id=uuid4(), title="Основное"),
                header="Комментарий",
            ),
        ),
        organizations=(SimpleNamespace(id=organization_id, name="Администрация", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
        import_mode="strict",
        work_experience_as_of_date=date(2026, 7, 17),
        title_header="Название карточки",
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки", "Комментарий"])
    sheet.append([1, "", "Нужны уточнения"])
    sheet.append([2, "Карточка для импорта", "Готово"])
    workbook.create_sheet("_registry_engine")["B1"] = "{}"
    output = BytesIO()
    workbook.save(output)
    created_display_names: list[object] = []

    class ImportSession:
        def begin_nested(self) -> object:
            return nullcontext()

    class ImportCardService:
        def __init__(self, _session: object) -> None:
            pass

        def validate_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

        def create_card_for_actor(self, **kwargs: object) -> object:
            created_display_names.append(kwargs["display_name"])
            return SimpleNamespace(id=uuid4())

        def set_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

    class AuditService:
        def __init__(self, _session: object) -> None:
            pass

        def record_user_event(self, **_kwargs: object) -> None:
            pass

    service = import_export.TabularCardExchangeService(ImportSession())
    monkeypatch.setattr(import_export, "CardService", ImportCardService)
    monkeypatch.setattr(import_export, "AuditService", AuditService)
    monkeypatch.setattr(service, "_configuration_from_metadata", lambda **_kwargs: configuration)

    preview = service.preview_import_xlsx_for_actor(
        actor_user_id=actor_user_id,
        registry_id=registry_id,
        xlsx_content=output.getvalue(),
    )

    assert preview["rows"][0]["errors"] == ["Название карточки: заполните название карточки."]
    assert preview["rows"][1]["display_name"] == "Карточка для импорта"
    with pytest.raises(import_export.TabularCardImportValidationError):
        service.commit_import_xlsx_for_actor(
            actor_user_id=actor_user_id,
            registry_id=registry_id,
            xlsx_content=output.getvalue(),
        )

    sheet.delete_rows(2)
    output = BytesIO()
    workbook.save(output)
    service.commit_import_xlsx_for_actor(
        actor_user_id=actor_user_id,
        registry_id=registry_id,
        xlsx_content=output.getvalue(),
    )

    assert created_display_names == ["Карточка для импорта"]


def test_tabular_xlsx_declares_user_facing_columns_and_supported_field_types() -> None:
    assert import_export.tabular_xlsx_fixed_headers(True) == (
        "№ п/п",
        "Название карточки",
        "Организация",
    )
    assert import_export.tabular_xlsx_fixed_headers(False) == ("№ п/п", "Название карточки")
    assert getattr(import_export, "TABULAR_XLSX_SUPPORTED_FIELD_TYPES", None) == {
        "text",
        "number",
        "date",
        "datetime",
        "bool",
        "select",
        "multi_select",
        "organization_ref",
        "org_unit_ref",
        "work_experience",
    }


def test_xlsx_configuration_rejects_an_inactive_template(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    actor_user_id = uuid4()
    registry_id = uuid4()
    expected_registry_id = registry_id
    template = SimpleNamespace(
        id=uuid4(),
        registry_id=registry_id,
        archived_at=None,
        is_active=False,
    )

    class AllowCardsManage:
        def __init__(self, _session: object) -> None:
            pass

        def has_permission(
            self,
            user_id: object,
            permission_code: object,
            *,
            registry_id: object,
        ) -> bool:
            assert (user_id, permission_code, registry_id) == (
                actor_user_id,
                "cards.manage",
                expected_registry_id,
            )
            return True

    class TemplateSession:
        def get(self, model: object, _model_id: object) -> object:
            assert model is import_export.CardTemplate
            return template

    monkeypatch.setattr(import_export, "PermissionService", AllowCardsManage)
    service = import_export.TabularCardExchangeService(TemplateSession())  # type: ignore[arg-type]

    with pytest.raises(import_export.ImportExportServiceError, match="Шаблон карточки"):
        service._configuration_for_actor(  # noqa: SLF001
            actor_user_id=actor_user_id,
            registry_id=registry_id,
            card_template_id=template.id,
            field_ids=[uuid4()],
            organization_ids=[uuid4()],
            include_organization_column=True,
            fixed_organization_id=None,
            require_fixed_organization=False,
        )


def test_tabular_xlsx_orders_fields_by_block_then_field_position() -> None:
    first_block = SimpleNamespace(id=uuid4(), position=1)
    second_block = SimpleNamespace(id=uuid4(), position=2)
    first_field = SimpleNamespace(id=uuid4(), position=1)
    second_field = SimpleNamespace(id=uuid4(), position=2)
    later_block_field = SimpleNamespace(id=uuid4(), position=0)

    ordered = import_export.TabularCardExchangeService._order_selected_fields(
        [
            (later_block_field, second_block),
            (second_field, first_block),
            (first_field, first_block),
        ]
    )

    assert [field.id for field, _block in ordered] == [
        first_field.id,
        second_field.id,
        later_block_field.id,
    ]


def test_tabular_xlsx_template_is_wide_formatted_and_contains_hidden_mapping() -> None:
    field = SimpleNamespace(
        id=uuid4(),
        label="Дата рождения",
        field_type="date",
    )
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=uuid4(),
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(
            import_export.TabularWorkbookField(
                field=field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Дата рождения",
            ),
        ),
        organizations=(SimpleNamespace(id=uuid4(), name="Администрация", code="admin"),),
        include_organization_column=True,
        fixed_organization_id=None,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
    )

    content = import_export.TabularCardExchangeService(MagicMock())._build_workbook(
        actor_user_id=uuid4(),
        configuration=configuration,
        cards=None,
    )

    workbook = load_workbook(filename=__import__("io").BytesIO(content), data_only=True)
    sheet = workbook["Карточки"]
    assert [cell.value for cell in sheet[1][:4]] == [
        "№ п/п",
        "Название карточки",
        "Организация",
        "Дата рождения",
    ]
    assert sheet["A2"].value == 1
    assert sheet["B2"].value is None
    assert sheet["C2"].value == "Администрация (admin)"
    assert sheet["D2"].number_format == 'DD"."MM"."YYYY'
    assert workbook["_registry_engine"].sheet_state == "hidden"


def test_tabular_xlsx_template_hides_organization_column_and_records_import_target() -> None:
    field = SimpleNamespace(id=uuid4(), label="Дата рождения", field_type="date")
    organization = SimpleNamespace(id=uuid4(), name="Администрация", code="admin")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=uuid4(),
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(
            import_export.TabularWorkbookField(
                field=field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Дата рождения",
            ),
        ),
        organizations=(organization,),
        include_organization_column=False,
        fixed_organization_id=organization.id,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
    )

    content = import_export.TabularCardExchangeService(MagicMock())._build_workbook(
        actor_user_id=uuid4(),
        configuration=configuration,
        cards=None,
    )

    workbook = load_workbook(filename=__import__("io").BytesIO(content), data_only=True)
    sheet = workbook["Карточки"]
    metadata = json.loads(workbook["_registry_engine"]["B1"].value)
    validations = {validation.sqref for validation in sheet.data_validations.dataValidation}

    assert [cell.value for cell in sheet[1][:3]] == ["№ п/п", "Название карточки", "Дата рождения"]
    assert sheet["A2"].value == 1
    assert sheet["B2"].value is None
    assert "C2:C101" not in validations
    assert metadata["include_organization_column"] is False
    assert metadata["fixed_organization_id"] == str(organization.id)


def test_tabular_xlsx_template_offers_reference_values_for_single_and_multiple_selects() -> None:
    select_field = SimpleNamespace(id=uuid4(), label="Статус", field_type="select")
    multi_select_field = SimpleNamespace(id=uuid4(), label="Категории", field_type="multi_select")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=uuid4(),
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(
            import_export.TabularWorkbookField(
                field=select_field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Статус",
            ),
            import_export.TabularWorkbookField(
                field=multi_select_field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Категории",
            ),
        ),
        organizations=(SimpleNamespace(id=uuid4(), name="Администрация", code="admin"),),
        include_organization_column=True,
        fixed_organization_id=None,
        organization_labels={},
        reference_labels={
            select_field.id: {"Новый": uuid4(), "Готово": uuid4()},
            multi_select_field.id: {"Основная": uuid4(), "Дополнительная": uuid4()},
        },
        unit_organization_ids={},
    )

    content = import_export.TabularCardExchangeService(MagicMock())._build_workbook(
        actor_user_id=uuid4(),
        configuration=configuration,
        cards=None,
    )

    workbook = load_workbook(filename=__import__("io").BytesIO(content), data_only=True)
    sheet = workbook["Карточки"]
    validations = {
        validation.sqref: validation for validation in sheet.data_validations.dataValidation
    }
    assert validations["C2:C101"].formula1 == "=organization_choices"
    assert validations["D2:D101"].formula1 == f"=field_{select_field.id.hex}_choices"
    assert validations["E2:E101"].formula1 == f"=field_{multi_select_field.id.hex}_choices"
    assert validations["E2:E101"].showErrorMessage is False


def test_tabular_xlsx_exports_readable_organization_and_unit_values(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    card_organization_id = uuid4()
    organization_field = SimpleNamespace(
        id=uuid4(),
        label="Организация для согласования",
        field_type="organization_ref",
    )
    org_unit_field = SimpleNamespace(
        id=uuid4(),
        label="Подразделение для согласования",
        field_type="org_unit_ref",
    )
    management_id = uuid4()
    department_id = uuid4()
    organization = SimpleNamespace(id=card_organization_id, name="Администрация", code="admin")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=uuid4(),
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(
            import_export.TabularWorkbookField(
                field=organization_field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Организация для согласования",
            ),
            import_export.TabularWorkbookField(
                field=org_unit_field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Подразделение для согласования",
            ),
        ),
        organizations=(organization,),
        include_organization_column=True,
        fixed_organization_id=None,
        organization_labels={"Администрация (admin)": card_organization_id},
        reference_labels={
            organization_field.id: {"Администрация": card_organization_id},
            org_unit_field.id: {"Администрация → Управление → Отдел": department_id},
        },
        unit_organization_ids={
            management_id: card_organization_id,
            department_id: card_organization_id,
        },
    )
    card = SimpleNamespace(id=uuid4(), organization_id=card_organization_id)

    class CardReader:
        def __init__(self, _session: object) -> None:
            pass

        def read_card_for_actor(self, *, actor_user_id: object, card_id: object) -> object:
            assert card_id == card.id
            return SimpleNamespace(
                blocks={
                    "main": SimpleNamespace(
                        instances=[
                            SimpleNamespace(
                                fields={
                                    "organization": SimpleNamespace(
                                        field_id=organization_field.id,
                                        value=card_organization_id,
                                    ),
                                    "unit": SimpleNamespace(
                                        field_id=org_unit_field.id,
                                        value=department_id,
                                    ),
                                }
                            )
                        ]
                    )
                }
            )

    monkeypatch.setattr(import_export, "CardService", CardReader)
    content = import_export.TabularCardExchangeService(MagicMock())._build_workbook(
        actor_user_id=uuid4(),
        configuration=configuration,
        cards=[card],
    )

    sheet = load_workbook(filename=BytesIO(content), data_only=True)["Карточки"]
    assert sheet["D2"].value == "Администрация"
    assert sheet["E2"].value == "Администрация → Управление → Отдел"
    assert str(card_organization_id) not in {sheet["D2"].value, sheet["E2"].value}
    assert str(department_id) not in {sheet["D2"].value, sheet["E2"].value}


def test_tabular_xlsx_round_trips_work_experience_as_three_columns_with_batch_as_of_date(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    actor_user_id = uuid4()
    registry_id = uuid4()
    organization_id = uuid4()
    field = SimpleNamespace(id=uuid4(), label="Стаж", field_type="work_experience")
    organization = SimpleNamespace(id=organization_id, name="Администрация", code="admin")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(
            import_export.TabularWorkbookField(
                field=field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Стаж",
            ),
        ),
        organizations=(organization,),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
        work_experience_as_of_date=date(2024, 6, 30),
    )
    card = SimpleNamespace(
        id=uuid4(),
        organization_id=organization_id,
        display_name="Карточка для проверки стажа",
    )

    class ExportCardService:
        def __init__(self, _session: object) -> None:
            pass

        def read_card_for_actor(self, *, actor_user_id: object, card_id: object) -> object:
            assert (actor_user_id, card_id) == (actor_user_id, card.id)
            return SimpleNamespace(
                blocks={
                    "main": SimpleNamespace(
                        instances=[
                            SimpleNamespace(
                                fields={
                                    "work_experience": SimpleNamespace(
                                        field_id=field.id,
                                        value={
                                            "days": 16,
                                            "months": 3,
                                            "years": 9,
                                        },
                                    )
                                }
                            )
                        ]
                    )
                }
            )

    monkeypatch.setattr(import_export, "CardService", ExportCardService)
    service = import_export.TabularCardExchangeService(MagicMock())
    content = service._build_workbook(
        actor_user_id=actor_user_id,
        configuration=configuration,
        cards=[card],
    )

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook["Карточки"]
    assert [cell.value for cell in sheet[1][:5]] == [
        "№ п/п",
        "Название карточки",
        "Стаж: дни",
        "Стаж: месяцы",
        "Стаж: годы",
    ]
    assert [sheet[cell].value for cell in ("C2", "D2", "E2")] == [16, 3, 9]
    assert [sheet[cell].number_format for cell in ("C2", "D2", "E2")] == ["0", "0", "0"]

    class ImportSession:
        def begin_nested(self) -> object:
            return nullcontext()

    set_value_calls: list[dict[str, object]] = []

    class ImportCardService:
        def __init__(self, _session: object) -> None:
            pass

        def create_card_for_actor(self, **_kwargs: object) -> object:
            return SimpleNamespace(id=uuid4())

        def set_field_value_for_actor(self, **kwargs: object) -> None:
            set_value_calls.append(kwargs)

        def validate_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

    class AuditService:
        def __init__(self, _session: object) -> None:
            pass

        def record_user_event(self, **_kwargs: object) -> None:
            pass

    monkeypatch.setattr(import_export, "CardService", ImportCardService)
    monkeypatch.setattr(import_export, "AuditService", AuditService)
    import_service = import_export.TabularCardExchangeService(ImportSession())
    monkeypatch.setattr(
        import_service,
        "_configuration_from_metadata",
        lambda **_kwargs: configuration,
    )

    preview = import_service.preview_import_xlsx_for_actor(
        actor_user_id=actor_user_id,
        registry_id=registry_id,
        xlsx_content=content,
    )
    committed = import_service.commit_import_xlsx_for_actor(
        actor_user_id=actor_user_id,
        registry_id=registry_id,
        xlsx_content=content,
    )

    assert preview["rows"][0]["values"] == {field.id: {"days": 16, "months": 3, "years": 9}}
    assert committed["summary"] == {
        "created_cards": 1,
        "field_values_written": 1,
        "created_reference_items": 0,
    }
    assert len(set_value_calls) == 1
    assert set_value_calls[0]["actor_user_id"] == actor_user_id
    assert set_value_calls[0]["field_id"] == field.id
    assert set_value_calls[0]["value"] == {"days": 16, "months": 3, "years": 9}
    assert set_value_calls[0]["work_experience_as_of_date"] == date(2024, 6, 30)


def test_tabular_xlsx_preview_and_commit_report_text_validation_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    actor_user_id = uuid4()
    registry_id = uuid4()
    organization_id = uuid4()
    field = SimpleNamespace(
        id=uuid4(),
        label="ФИО",
        field_type="text",
        validation_json={
            "kind": "russian_text",
            "message": "Введите ФИО русскими буквами",
        },
    )
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(
            import_export.TabularWorkbookField(
                field=field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="ФИО",
            ),
        ),
        organizations=(SimpleNamespace(id=organization_id, name="Администрация", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки", "ФИО"])
    sheet.append([1, "Проверка ФИО", "Иванов 7"])
    workbook.create_sheet("_registry_engine")["B1"] = "{}"
    content = BytesIO()
    workbook.save(content)
    validation_calls: list[dict[str, object]] = []

    class ImportCardService:
        def __init__(self, _session: object) -> None:
            pass

        def validate_field_value_for_actor(self, **kwargs: object) -> None:
            validation_calls.append(kwargs)
            raise InvalidFieldValueError("Введите ФИО русскими буквами")

    service = import_export.TabularCardExchangeService(MagicMock())
    monkeypatch.setattr(import_export, "CardService", ImportCardService)
    monkeypatch.setattr(
        service,
        "_configuration_from_metadata",
        lambda **_kwargs: configuration,
    )

    preview = service.preview_import_xlsx_for_actor(
        actor_user_id=actor_user_id,
        registry_id=registry_id,
        xlsx_content=content.getvalue(),
    )

    assert preview["rows"][0]["status"] == "invalid"
    assert preview["rows"][0]["errors"] == ["ФИО: Введите ФИО русскими буквами"]
    assert validation_calls == [
        {
            "actor_user_id": actor_user_id,
            "registry_id": registry_id,
            "organization_id": organization_id,
            "field_id": field.id,
            "value": "Иванов 7",
        }
    ]
    with pytest.raises(import_export.TabularCardImportValidationError) as exc_info:
        service.commit_import_xlsx_for_actor(
            actor_user_id=actor_user_id,
            registry_id=registry_id,
            xlsx_content=content.getvalue(),
        )
    assert exc_info.value.preview["rows"][0]["errors"] == ["ФИО: Введите ФИО русскими буквами"]


def test_tabular_xlsx_requires_all_or_none_work_experience_columns(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry_id = uuid4()
    organization_id = uuid4()
    field = SimpleNamespace(id=uuid4(), label="Стаж", field_type="work_experience")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(
            import_export.TabularWorkbookField(
                field=field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Стаж",
            ),
        ),
        organizations=(SimpleNamespace(id=organization_id, name="Администрация", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки", "Стаж: дни", "Стаж: месяцы", "Стаж: годы"])
    sheet.append([1, "Без стажа", None, None, None])
    sheet.append([2, "Неполный стаж", 16, None, 9])
    sheet.append([3, "Полный стаж", 16, 3, 9])
    metadata_sheet = workbook.create_sheet("_registry_engine")
    metadata_sheet["B1"] = "{}"
    content = BytesIO()
    workbook.save(content)

    service = import_export.TabularCardExchangeService(MagicMock())

    class ImportCardService:
        def __init__(self, _session: object) -> None:
            pass

        def validate_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

    monkeypatch.setattr(import_export, "CardService", ImportCardService)
    monkeypatch.setattr(
        service,
        "_configuration_from_metadata",
        lambda **_kwargs: configuration,
    )

    preview = service.preview_import_xlsx_for_actor(
        actor_user_id=uuid4(),
        registry_id=registry_id,
        xlsx_content=content.getvalue(),
    )

    assert preview["rows"][0]["status"] == "valid"
    assert preview["rows"][0]["values"] == {}
    assert preview["rows"][1]["status"] == "invalid"
    assert preview["rows"][1]["errors"] == [
        "Стаж: заполните дни, месяцы и годы стажа либо оставьте все три значения пустыми."
    ]
    assert preview["rows"][2]["status"] == "valid"
    assert preview["rows"][2]["values"] == {field.id: {"days": 16, "months": 3, "years": 9}}


def test_tabular_xlsx_rejects_formulas_in_visible_data_cells(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry_id = uuid4()
    organization_id = uuid4()
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(),
        organizations=(SimpleNamespace(id=organization_id, name="Администрация", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки"])
    sheet.append([1, "=1+1"])
    workbook.create_sheet("_registry_engine")["B1"] = "{}"
    content = BytesIO()
    workbook.save(content)

    service = import_export.TabularCardExchangeService(MagicMock())
    monkeypatch.setattr(service, "_configuration_from_metadata", lambda **_kwargs: configuration)

    with pytest.raises(import_export.ImportExportServiceError, match="формул"):
        service.preview_import_xlsx_for_actor(
            actor_user_id=uuid4(),
            registry_id=registry_id,
            xlsx_content=content.getvalue(),
        )


def test_tabular_xlsx_export_escapes_formula_leading_text_and_marks_it_as_text(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    field = SimpleNamespace(id=uuid4(), label="Комментарий", field_type="text")
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=uuid4(),
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(
            import_export.TabularWorkbookField(
                field=field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Комментарий",
            ),
        ),
        organizations=(SimpleNamespace(id=uuid4(), name="Администрация", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=uuid4(),
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
    )
    card = SimpleNamespace(
        id=uuid4(),
        organization_id=configuration.fixed_organization_id,
        display_name='=HYPERLINK("https://example.test")',
    )

    class CardReader:
        def __init__(self, _session: object) -> None:
            pass

        def read_card_for_actor(self, **_kwargs: object) -> object:
            return SimpleNamespace(
                blocks={
                    "main": SimpleNamespace(
                        instances=[
                            SimpleNamespace(
                                fields={
                                    "comment": SimpleNamespace(
                                        field_id=field.id,
                                        value="=2+2",
                                    )
                                }
                            )
                        ]
                    )
                }
            )

    monkeypatch.setattr(import_export, "CardService", CardReader)
    content = import_export.TabularCardExchangeService(MagicMock())._build_workbook(
        actor_user_id=uuid4(),
        configuration=configuration,
        cards=[card],
    )

    sheet = load_workbook(filename=BytesIO(content), data_only=False)["Карточки"]
    assert sheet["B2"].value == '\'=HYPERLINK("https://example.test")'
    assert sheet["C2"].value == "'=2+2"
    assert sheet["B2"].number_format == "@"
    assert sheet["C2"].number_format == "@"


@pytest.mark.parametrize(
    ("setting_name", "setting_value", "workbook_setup"),
    [
        ("max_import_sheets", 2, lambda workbook, sheet: workbook.create_sheet("Лишний")),
        (
            "max_import_columns",
            2,
            lambda _workbook, sheet: sheet.cell(row=1, column=3, value="Лишняя"),
        ),
        (
            "max_import_cells",
            2,
            lambda _workbook, sheet: sheet.cell(row=2, column=2, value="Название"),
        ),
        ("max_import_uncompressed_bytes", 1, lambda _workbook, _sheet: None),
    ],
)
def test_tabular_xlsx_rejects_configured_workbook_limits_before_row_parsing(
    monkeypatch: pytest.MonkeyPatch,
    setting_name: str,
    setting_value: int,
    workbook_setup: object,
) -> None:
    registry_id = uuid4()
    organization_id = uuid4()
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(),
        organizations=(SimpleNamespace(id=organization_id, name="Администрация", code="admin"),),
        include_organization_column=False,
        fixed_organization_id=organization_id,
        organization_labels={},
        reference_labels={},
        unit_organization_ids={},
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(["№ п/п", "Название карточки"])
    workbook.create_sheet("_registry_engine")["B1"] = "{}"
    workbook_setup(workbook, sheet)
    content = BytesIO()
    workbook.save(content)
    settings = SimpleNamespace(
        max_import_rows=10_000,
        max_import_sheets=8,
        max_import_columns=200,
        max_import_cells=500_000,
        max_import_uncompressed_bytes=50 * 1024 * 1024,
    )
    setattr(settings, setting_name, setting_value)

    service = import_export.TabularCardExchangeService(MagicMock())
    monkeypatch.setattr(import_export, "get_settings", lambda: settings)
    monkeypatch.setattr(service, "_configuration_from_metadata", lambda **_kwargs: configuration)

    with pytest.raises(import_export.ImportExportServiceError, match="лимит"):
        service.preview_import_xlsx_for_actor(
            actor_user_id=uuid4(),
            registry_id=registry_id,
            xlsx_content=content.getvalue(),
        )


def test_tabular_xlsx_preview_resolves_organization_references_and_rejects_foreign_unit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    registry_id = uuid4()
    card_organization_id = uuid4()
    foreign_organization_id = uuid4()
    organization_field = SimpleNamespace(id=uuid4(), field_type="organization_ref")
    org_unit_field = SimpleNamespace(id=uuid4(), field_type="org_unit_ref")
    local_unit_id = uuid4()
    foreign_unit_id = uuid4()
    configuration = import_export.TabularWorkbookConfiguration(
        registry_id=registry_id,
        template=SimpleNamespace(id=uuid4(), name="Сведения"),
        fields=(
            import_export.TabularWorkbookField(
                field=organization_field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Организация для согласования",
            ),
            import_export.TabularWorkbookField(
                field=org_unit_field,
                block=SimpleNamespace(id=uuid4(), title="Основные сведения"),
                header="Подразделение для согласования",
            ),
        ),
        organizations=(
            SimpleNamespace(id=card_organization_id, name="Администрация", code="admin"),
            SimpleNamespace(id=foreign_organization_id, name="Школа", code="school"),
        ),
        include_organization_column=True,
        fixed_organization_id=None,
        organization_labels={
            "Администрация (admin)": card_organization_id,
            "Школа (school)": foreign_organization_id,
        },
        reference_labels={
            organization_field.id: {"Администрация": card_organization_id},
            org_unit_field.id: {
                "Администрация → Управление → Отдел": local_unit_id,
                "Школа → Отдел": foreign_unit_id,
            },
        },
        unit_organization_ids={
            local_unit_id: card_organization_id,
            foreign_unit_id: foreign_organization_id,
        },
    )

    class ImportCardService:
        def __init__(self, _session: object) -> None:
            pass

        def validate_field_value_for_actor(self, **_kwargs: object) -> None:
            pass

    service = import_export.TabularCardExchangeService(MagicMock())
    monkeypatch.setattr(import_export, "CardService", ImportCardService)
    monkeypatch.setattr(
        service,
        "_configuration_from_metadata",
        lambda **_kwargs: configuration,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Карточки"
    sheet.append(
        [
            "№ п/п",
            "Название карточки",
            "Организация",
            "Организация для согласования",
            "Подразделение для согласования",
        ]
    )
    sheet.append(
        [
            1,
            "Карточка администрации",
            "Администрация (admin)",
            "Администрация",
            "Администрация → Управление → Отдел",
        ]
    )
    sheet.append(
        [
            2,
            "Карточка с неверным подразделением",
            "Администрация (admin)",
            "Администрация",
            "Школа → Отдел",
        ]
    )
    metadata_sheet = workbook.create_sheet("_registry_engine")
    metadata_sheet["B1"] = "{}"
    content = BytesIO()
    workbook.save(content)

    preview = service.preview_import_xlsx_for_actor(
        actor_user_id=uuid4(),
        registry_id=registry_id,
        xlsx_content=content.getvalue(),
    )

    assert preview["rows"][0]["status"] == "valid"
    assert preview["rows"][0]["values"] == {
        organization_field.id: card_organization_id,
        org_unit_field.id: local_unit_id,
    }
    assert preview["rows"][1]["status"] == "invalid"
    assert preview["rows"][1]["errors"] == [
        "Подразделение для согласования: подразделение недоступно выбранной организации."
    ]


def test_xlsx_download_headers_are_safe_for_http_response_encoding() -> None:
    headers = import_export_endpoint.xlsx_download_headers("registry-cards.xlsx")

    response = Response(content=b"xlsx", headers=headers)

    assert response.headers["x-document-filename"] == "registry-cards.xlsx"
    assert response.headers["content-disposition"] == 'attachment; filename="registry-cards.xlsx"'
