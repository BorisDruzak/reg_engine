import io
import os
from collections.abc import Iterator
from pathlib import Path
from uuid import UUID

import pytest
from alembic import command
from alembic.config import Config
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select, text
from sqlalchemy.engine import Engine, make_url
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.main import create_app
from app.models import AccessGrant, Card, Permission, ReferenceItem, Role, User, role_permissions
from app.services.cards import CardService
from app.services.organizations import OrganizationService
from app.services.references import ReferenceListService
from app.services.registry_schema import RegistrySchemaService


def _require_test_database_url() -> str:
    database_url = os.environ.get("TEST_DATABASE_URL")
    if not database_url:
        pytest.skip("TEST_DATABASE_URL is required for disposable PostgreSQL API tests.")
    database_name = make_url(database_url).database or ""
    if database_name == "reg_engine" or not database_name.endswith("_test"):
        pytest.fail("TEST_DATABASE_URL must point to a disposable database ending with '_test'.")
    return database_url


def _alembic_config() -> Config:
    backend_root = Path(__file__).resolve().parents[1]
    config = Config(str(backend_root / "alembic.ini"))
    config.set_main_option("script_location", str(backend_root / "migrations"))
    return config


@pytest.fixture(scope="module")
def migrated_test_engine() -> Iterator[Engine]:
    database_url = _require_test_database_url()
    engine = create_engine(database_url)
    with engine.connect().execution_options(isolation_level="AUTOCOMMIT") as connection:
        connection.execute(text("DROP SCHEMA IF EXISTS public CASCADE"))
        connection.execute(text("CREATE SCHEMA public"))
    previous_url = os.environ.get("TEST_DATABASE_URL")
    os.environ["TEST_DATABASE_URL"] = database_url
    try:
        command.upgrade(_alembic_config(), "head")
    finally:
        if previous_url is None:
            os.environ.pop("TEST_DATABASE_URL", None)
        else:
            os.environ["TEST_DATABASE_URL"] = previous_url
    try:
        yield engine
    finally:
        engine.dispose()


@pytest.fixture()
def db_session(migrated_test_engine: Engine) -> Iterator[Session]:
    connection = migrated_test_engine.connect()
    transaction = connection.begin()
    session = Session(bind=connection, expire_on_commit=False)
    try:
        yield session
    finally:
        session.close()
        transaction.rollback()
        connection.close()


@pytest.fixture()
def api_client(db_session: Session) -> Iterator[TestClient]:
    from app.api.dependencies import get_db_session

    previous_allow_dev_actor = os.environ.get("ALLOW_DEV_ACTOR_HEADER")
    os.environ["ALLOW_DEV_ACTOR_HEADER"] = "true"
    get_settings.cache_clear()
    app = create_app()

    def override_session() -> Iterator[Session]:
        yield db_session

    app.dependency_overrides[get_db_session] = override_session
    try:
        with TestClient(app) as client:
            yield client
    finally:
        app.dependency_overrides.clear()
        if previous_allow_dev_actor is None:
            os.environ.pop("ALLOW_DEV_ACTOR_HEADER", None)
        else:
            os.environ["ALLOW_DEV_ACTOR_HEADER"] = previous_allow_dev_actor
        get_settings.cache_clear()


def _actor_headers(user_id: UUID) -> dict[str, str]:
    return {"X-Actor-User-Id": str(user_id)}


def _context(session: Session) -> dict[str, object]:
    system = User(
        email="xlsx-system@example.test",
        display_name="Системный администратор",
        is_superuser=True,
    )
    scoped = User(email="xlsx-scoped@example.test", display_name="Администратор организации")
    session.add_all([system, scoped])
    session.flush()

    role = Role(code="xlsx_cards_manage", name="Администратор XLSX")
    permission = Permission(code="cards.manage", description="Управление карточками")
    session.add_all([role, permission])
    session.flush()
    session.execute(role_permissions.insert().values(role_id=role.id, permission_id=permission.id))

    organizations = OrganizationService(session)
    root = organizations.create_root_for_actor(
        actor_user_id=system.id,
        code="xlsx-root",
        name="Корневая организация",
    )
    child = organizations.create_child(
        parent_id=root.id,
        code="xlsx-child",
        name="Доступная организация",
        created_by=system.id,
    )
    sibling = organizations.create_child(
        parent_id=root.id,
        code="xlsx-sibling",
        name="Недоступная организация",
        created_by=system.id,
    )
    session.add(
        AccessGrant(
            user_id=scoped.id,
            role_id=role.id,
            organization_id=child.id,
            include_descendants=True,
            created_by=system.id,
        )
    )

    schema = RegistrySchemaService(session)
    registry = schema.create_registry_for_actor(
        actor_user_id=system.id,
        code="xlsx-registry",
        name="Реестр XLSX",
    )
    block = schema.create_block_for_actor(
        actor_user_id=system.id,
        registry_id=registry.id,
        code="main",
        title="Основные сведения",
    )
    status = schema.create_field_for_actor(
        actor_user_id=system.id,
        block_id=block.id,
        code="status",
        label="Статус",
        field_type="text",
    )
    template = schema.ensure_base_card_template_for_registry(
        registry_id=registry.id,
        actor_user_id=system.id,
    )
    source_card = CardService(session).create_card_for_actor(
        actor_user_id=scoped.id,
        registry_id=registry.id,
        organization_id=child.id,
        card_template_id=template.id,
    )
    CardService(session).set_field_value_for_actor(
        actor_user_id=scoped.id,
        card_id=source_card.id,
        field_id=status.id,
        value="Готово",
    )
    session.flush()
    return {
        "system": system,
        "scoped": scoped,
        "registry": registry,
        "child": child,
        "sibling": sibling,
        "template": template,
        "status": status,
        "source_card": source_card,
    }


def _selection(context: dict[str, object]) -> dict[str, object]:
    return {
        "card_template_id": str(context["template"].id),
        "field_ids": [str(context["status"].id)],
        "organization_ids": [str(context["child"].id)],
        "include_organization_column": True,
    }


def _url(context: dict[str, object], suffix: str) -> str:
    return f"/api/v1/registries/{context['registry'].id}/tabular-xlsx-card-exchange/{suffix}"


def _add_global_reference_select_field(
    session: Session,
    context: dict[str, object],
) -> dict[str, object]:
    reference_service = ReferenceListService(session)
    reference_list = reference_service.create_reference_list_for_actor(
        actor_user_id=context["system"].id,
        registry_id=context["registry"].id,
        code="xlsx-statuses",
        name="Статусы XLSX",
    )
    existing_item = reference_service.create_reference_item_for_actor(
        actor_user_id=context["system"].id,
        list_id=reference_list.id,
        code="existing",
        label="Существующее значение",
    )
    schema = RegistrySchemaService(session)
    field = schema.create_field_for_actor(
        actor_user_id=context["system"].id,
        block_id=context["status"].block_id,
        code="xlsx_status",
        label="Статус XLSX",
        field_type="select",
        options_source_type="reference_list",
        options_source_id=reference_list.id,
    )
    context["template"] = schema.ensure_base_card_template_for_registry(
        registry_id=context["registry"].id,
        actor_user_id=context["system"].id,
    )
    return {"field": field, "reference_list": reference_list, "existing_item": existing_item}


def _grant_reference_edit_permission(session: Session, context: dict[str, object]) -> None:
    role = Role(code="xlsx_reference_editor", name="Редактор справочника XLSX")
    permission = Permission(
        code="registry.schema.manage",
        description="Управление схемой реестра",
    )
    session.add_all([role, permission])
    session.flush()
    session.execute(role_permissions.insert().values(role_id=role.id, permission_id=permission.id))
    session.add(
        AccessGrant(
            user_id=context["scoped"].id,
            role_id=role.id,
            registry_id=context["registry"].id,
            created_by=context["system"].id,
        )
    )
    session.flush()


def _reference_selection(
    context: dict[str, object],
    reference_context: dict[str, object],
    *,
    import_mode: str,
) -> dict[str, object]:
    return {
        "card_template_id": str(context["template"].id),
        "field_ids": [str(reference_context["field"].id)],
        "organization_ids": [str(context["child"].id)],
        "include_organization_column": False,
        "fixed_organization_id": str(context["child"].id),
        "import_mode": import_mode,
    }


def _upload_xlsx(
    api_client: TestClient,
    *,
    url: str,
    content: bytes,
    headers: dict[str, str],
):
    return api_client.post(
        url,
        files={
            "file": (
                "template.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )


def test_tabular_xlsx_options_are_scoped_to_card_management(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _context(db_session)

    response = api_client.get(
        _url(context, "options"),
        headers=_actor_headers(context["scoped"].id),
    )

    assert response.status_code == 200, response.text
    payload = response.json()
    assert [item["id"] for item in payload["organizations"]] == [str(context["child"].id)]
    template = next(
        item for item in payload["templates"] if item["id"] == str(context["template"].id)
    )
    assert template["fields"] == [
        {
            "id": str(context["status"].id),
            "label": "Статус",
            "block_title": "Основные сведения",
            "field_type": "text",
            "supported": True,
            "unsupported_reason": None,
        }
    ]


def test_tabular_xlsx_export_and_template_are_wide_and_readable(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _context(db_session)
    headers = _actor_headers(context["scoped"].id)
    selection = _selection(context)

    exported = api_client.post(_url(context, "export"), json=selection, headers=headers)
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(io.BytesIO(exported.content), data_only=True)
    sheet = workbook["Карточки"]
    assert [cell.value for cell in sheet[1][:4]] == [
        "№ п/п",
        "Название карточки",
        "Организация",
        "Статус",
    ]
    assert sheet["A2"].value == 1
    assert sheet["B2"].value == context["template"].name
    assert sheet["D2"].value == "Готово"
    assert workbook["_registry_engine"].sheet_state == "hidden"

    template_response = api_client.post(
        _url(context, "import-template"),
        json=selection,
        headers=headers,
    )
    assert template_response.status_code == 200, template_response.text
    template_workbook = load_workbook(io.BytesIO(template_response.content), data_only=True)
    template_sheet = template_workbook["Карточки"]
    assert template_sheet["A2"].value == 1
    assert template_sheet["B2"].value is None
    assert template_sheet["C2"].value == "Доступная организация (xlsx-child)"
    assert template_sheet["D2"].value is None


def test_tabular_xlsx_import_previews_and_creates_cards_atomically(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _context(db_session)
    headers = _actor_headers(context["scoped"].id)
    template_response = api_client.post(
        _url(context, "import-template"),
        json=_selection(context),
        headers=headers,
    )
    workbook = load_workbook(io.BytesIO(template_response.content))
    workbook["Карточки"]["B2"] = "Импортированная карточка"
    workbook["Карточки"]["D2"] = "Импортировано"
    content = io.BytesIO()
    workbook.save(content)

    preview = api_client.post(
        _url(context, "import/preview"),
        files={
            "file": (
                "template.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["summary"] == {
        "total_rows": 1,
        "valid_rows": 1,
        "invalid_rows": 0,
        "would_create_cards": 1,
        "would_create_reference_items": 0,
    }

    committed = api_client.post(
        _url(context, "import/commit"),
        files={
            "file": (
                "template.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )
    assert committed.status_code == 200, committed.text
    assert committed.json()["summary"] == {
        "created_cards": 1,
        "field_values_written": 1,
        "created_reference_items": 0,
    }
    created = db_session.scalar(
        select(Card)
        .where(Card.display_name == "Импортированная карточка")
        .order_by(Card.created_at.desc())
    )
    assert created is not None
    assert created.card_template_id == context["template"].id
    values = CardService(db_session).read_card_for_actor(
        actor_user_id=context["scoped"].id,
        card_id=created.id,
    )
    assert values.blocks["main"].instances[0].fields["status"].value == "Импортировано"


def test_tabular_xlsx_strict_mode_creates_cards_only_for_existing_global_references(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _context(db_session)
    reference_context = _add_global_reference_select_field(db_session, context)
    headers = _actor_headers(context["scoped"].id)
    template_response = api_client.post(
        _url(context, "import-template"),
        json=_reference_selection(context, reference_context, import_mode="strict"),
        headers=headers,
    )
    assert template_response.status_code == 200, template_response.text
    workbook = load_workbook(io.BytesIO(template_response.content))
    sheet = workbook["Карточки"]
    sheet["B2"] = "Строгая карточка"
    sheet["C2"] = "Существующее значение"
    content = io.BytesIO()
    workbook.save(content)

    preview = _upload_xlsx(
        api_client,
        url=_url(context, "import/preview"),
        content=content.getvalue(),
        headers=headers,
    )

    assert preview.status_code == 200, preview.text
    assert preview.json()["summary"] == {
        "total_rows": 1,
        "valid_rows": 1,
        "invalid_rows": 0,
        "would_create_cards": 1,
        "would_create_reference_items": 0,
    }
    assert preview.json()["new_reference_items"] == []

    committed = _upload_xlsx(
        api_client,
        url=_url(context, "import/commit"),
        content=content.getvalue(),
        headers=headers,
    )

    assert committed.status_code == 200, committed.text
    assert committed.json()["summary"] == {
        "created_cards": 1,
        "field_values_written": 1,
        "created_reference_items": 0,
    }
    assert (
        db_session.scalar(
            select(func.count())
            .select_from(ReferenceItem)
            .where(ReferenceItem.list_id == reference_context["reference_list"].id)
        )
        == 1
    )
    assert (
        db_session.scalar(select(Card).where(Card.display_name == "Строгая карточка")) is not None
    )


def test_tabular_xlsx_enrichment_creates_exactly_previewed_global_references_and_cards(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _context(db_session)
    reference_context = _add_global_reference_select_field(db_session, context)
    _grant_reference_edit_permission(db_session, context)
    headers = _actor_headers(context["scoped"].id)
    template_response = api_client.post(
        _url(context, "import-template"),
        json=_reference_selection(
            context,
            reference_context,
            import_mode="enrich_global_references",
        ),
        headers=headers,
    )
    assert template_response.status_code == 200, template_response.text
    workbook = load_workbook(io.BytesIO(template_response.content))
    sheet = workbook["Карточки"]
    sheet["B2"] = "Карточка с новым статусом"
    sheet["C2"] = "Новый статус"
    sheet["B3"] = "Вторая карточка с новым статусом"
    sheet["C3"] = "  новый\u00a0  статус "
    content = io.BytesIO()
    workbook.save(content)

    preview = _upload_xlsx(
        api_client,
        url=_url(context, "import/preview"),
        content=content.getvalue(),
        headers=headers,
    )

    assert preview.status_code == 200, preview.text
    assert preview.json()["summary"] == {
        "total_rows": 2,
        "valid_rows": 2,
        "invalid_rows": 0,
        "would_create_cards": 2,
        "would_create_reference_items": 1,
    }
    assert preview.json()["new_reference_items"] == [
        {"field_label": "Статус XLSX", "label": "Новый статус"}
    ]

    committed = _upload_xlsx(
        api_client,
        url=_url(context, "import/commit"),
        content=content.getvalue(),
        headers=headers,
    )

    assert committed.status_code == 200, committed.text
    assert committed.json()["summary"] == {
        "created_cards": 2,
        "field_values_written": 2,
        "created_reference_items": 1,
    }
    assert (
        db_session.scalar(
            select(func.count())
            .select_from(ReferenceItem)
            .where(
                ReferenceItem.list_id == reference_context["reference_list"].id,
                ReferenceItem.label == "Новый статус",
            )
        )
        == 1
    )
    assert (
        db_session.scalar(
            select(func.count())
            .select_from(Card)
            .where(
                Card.display_name.in_(
                    ["Карточка с новым статусом", "Вторая карточка с новым статусом"]
                )
            )
        )
        == 2
    )


def test_tabular_xlsx_enrichment_without_reference_edit_permission_returns_safe_403(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _context(db_session)
    reference_context = _add_global_reference_select_field(db_session, context)
    headers = _actor_headers(context["scoped"].id)
    template_response = api_client.post(
        _url(context, "import-template"),
        json=_reference_selection(
            context,
            reference_context,
            import_mode="enrich_global_references",
        ),
        headers=headers,
    )
    assert template_response.status_code == 200, template_response.text
    workbook = load_workbook(io.BytesIO(template_response.content))
    sheet = workbook["Карточки"]
    sheet["B2"] = "Карточка без права"
    sheet["C2"] = "Новый статус без права"
    content = io.BytesIO()
    workbook.save(content)

    response = _upload_xlsx(
        api_client,
        url=_url(context, "import/preview"),
        content=content.getvalue(),
        headers=headers,
    )

    assert response.status_code == 403
    assert response.json() == {"detail": "Недостаточно прав для выполнения операции."}


def test_tabular_xlsx_invalid_enrichment_commit_leaves_no_cards_or_reference_items(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _context(db_session)
    reference_context = _add_global_reference_select_field(db_session, context)
    _grant_reference_edit_permission(db_session, context)
    headers = _actor_headers(context["scoped"].id)
    template_response = api_client.post(
        _url(context, "import-template"),
        json=_reference_selection(
            context,
            reference_context,
            import_mode="enrich_global_references",
        ),
        headers=headers,
    )
    assert template_response.status_code == 200, template_response.text
    workbook = load_workbook(io.BytesIO(template_response.content))
    sheet = workbook["Карточки"]
    sheet["B2"] = "Атомарная карточка"
    sheet["C2"] = "Атомарный новый статус"
    sheet["C3"] = "Атомарный новый статус"
    content = io.BytesIO()
    workbook.save(content)

    committed = _upload_xlsx(
        api_client,
        url=_url(context, "import/commit"),
        content=content.getvalue(),
        headers=headers,
    )

    assert committed.status_code == 400, committed.text
    assert committed.json()["detail"]["summary"]["invalid_rows"] == 1
    assert (
        db_session.scalar(
            select(func.count()).select_from(Card).where(Card.display_name == "Атомарная карточка")
        )
        == 0
    )
    assert (
        db_session.scalar(
            select(func.count())
            .select_from(ReferenceItem)
            .where(
                ReferenceItem.list_id == reference_context["reference_list"].id,
                ReferenceItem.label == "Атомарный новый статус",
            )
        )
        == 0
    )


def test_technical_card_exchange_routes_are_not_exposed(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _context(db_session)
    headers = _actor_headers(context["scoped"].id)
    registry_id = context["registry"].id

    assert (
        api_client.get(
            f"/api/v1/registries/{registry_id}/exports/cards?format=xlsx",
            headers=headers,
        ).status_code
        == 404
    )
    assert (
        api_client.post(
            f"/api/v1/registries/{registry_id}/imports/cards/preview",
            json={"csv_content": "legacy"},
            headers=headers,
        ).status_code
        == 404
    )
