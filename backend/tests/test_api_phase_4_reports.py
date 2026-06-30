import csv
import json
import os
from collections.abc import Iterator
from datetime import UTC, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

import pytest
from alembic import command
from alembic.config import Config
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import create_engine, select, text
from sqlalchemy.engine import Engine, make_url
from sqlalchemy.orm import Session

from app.api.dependencies import get_db_session
from app.core.config import get_settings
from app.main import create_app
from app.models import (
    AccessGrant,
    AuditEvent,
    Permission,
    ReportRun,
    ReportTemplate,
    Role,
    StoredFile,
    User,
    role_permissions,
)
from app.services.cards import CardService
from app.services.organizations import OrganizationService
from app.services.registry_schema import RegistrySchemaService
from app.services.reports import ReportService, _RenderedReport


def _require_test_database_url() -> str:
    database_url = os.environ.get("TEST_DATABASE_URL")
    if not database_url:
        pytest.skip("TEST_DATABASE_URL is required for disposable PostgreSQL report API tests.")

    database_name = make_url(database_url).database or ""
    if database_name == "reg_engine" or not database_name.endswith("_test"):
        pytest.fail("TEST_DATABASE_URL must point to a disposable database ending with '_test'.")

    return database_url


def _alembic_config() -> Config:
    backend_root = Path(__file__).resolve().parents[1]
    config = Config(str(backend_root / "alembic.ini"))
    config.set_main_option("script_location", str(backend_root / "migrations"))
    return config


def _run_alembic_upgrade(database_url: str) -> None:
    previous_url = os.environ.get("TEST_DATABASE_URL")
    os.environ["TEST_DATABASE_URL"] = database_url
    try:
        command.upgrade(_alembic_config(), "head")
    finally:
        if previous_url is None:
            os.environ.pop("TEST_DATABASE_URL", None)
        else:
            os.environ["TEST_DATABASE_URL"] = previous_url


def _reset_public_schema(engine: Engine) -> None:
    with engine.connect().execution_options(isolation_level="AUTOCOMMIT") as connection:
        connection.execute(text("DROP SCHEMA IF EXISTS public CASCADE"))
        connection.execute(text("CREATE SCHEMA public"))


@pytest.fixture(scope="module")
def migrated_test_engine() -> Iterator[Engine]:
    database_url = _require_test_database_url()
    engine = create_engine(database_url)

    _reset_public_schema(engine)
    _run_alembic_upgrade(database_url)

    try:
        yield engine
    finally:
        engine.dispose()


@pytest.fixture()
def db_session(migrated_test_engine: Engine) -> Iterator[Session]:
    connection = migrated_test_engine.connect()
    transaction = connection.begin()
    session = Session(bind=connection, expire_on_commit=False)

    try:
        yield session
    finally:
        session.close()
        transaction.rollback()
        connection.close()


@pytest.fixture()
def api_client(db_session: Session, tmp_path: Path) -> Iterator[TestClient]:
    previous_allow_dev_actor = os.environ.get("ALLOW_DEV_ACTOR_HEADER")
    previous_storage_root = os.environ.get("REG_ENGINE_STORAGE_ROOT")
    os.environ["ALLOW_DEV_ACTOR_HEADER"] = "true"
    os.environ["REG_ENGINE_STORAGE_ROOT"] = str(tmp_path)
    get_settings.cache_clear()
    app = create_app()

    def override_session() -> Iterator[Session]:
        yield db_session

    app.dependency_overrides[get_db_session] = override_session
    try:
        with TestClient(app) as client:
            yield client
    finally:
        _restore_env("ALLOW_DEV_ACTOR_HEADER", previous_allow_dev_actor)
        _restore_env("REG_ENGINE_STORAGE_ROOT", previous_storage_root)
        get_settings.cache_clear()


def _restore_env(name: str, value: str | None) -> None:
    if value is None:
        os.environ.pop(name, None)
    else:
        os.environ[name] = value


def _actor_headers(user_id: UUID) -> dict[str, str]:
    return {"X-Actor-User-Id": str(user_id)}


def _create_user(session: Session, email: str, *, is_superuser: bool = False) -> User:
    user = User(email=email, display_name=email, is_superuser=is_superuser)
    session.add(user)
    session.flush()
    return user


def _create_role_with_permissions(
    session: Session,
    role_code: str,
    permission_codes: list[str],
) -> Role:
    role = Role(code=role_code, name=role_code)
    session.add(role)
    session.flush()

    for permission_code in permission_codes:
        permission = Permission(code=permission_code, description=permission_code)
        session.add(permission)
        session.flush()
        session.execute(
            role_permissions.insert().values(role_id=role.id, permission_id=permission.id)
        )

    session.flush()
    return role


def _grant_access(
    session: Session,
    *,
    user_id: UUID,
    role_id: UUID,
    organization_id: UUID | None = None,
    registry_id: UUID | None = None,
    include_descendants: bool = True,
    created_by: UUID | None = None,
) -> AccessGrant:
    grant = AccessGrant(
        user_id=user_id,
        role_id=role_id,
        organization_id=organization_id,
        registry_id=registry_id,
        include_descendants=include_descendants,
        created_by=created_by,
    )
    session.add(grant)
    session.flush()
    return grant


def _report_api_context(db_session: Session) -> dict[str, Any]:
    system_admin = _create_user(db_session, "api-reports-system@example.test", is_superuser=True)
    schema_admin = _create_user(db_session, "api-reports-schema-admin@example.test")
    card_admin = _create_user(db_session, "api-reports-card-admin@example.test")
    sibling_admin = _create_user(db_session, "api-reports-sibling-admin@example.test")
    schema_role = _create_role_with_permissions(
        db_session,
        "api_reports_schema_admin",
        ["registry.schema.manage"],
    )
    card_role = _create_role_with_permissions(
        db_session,
        "api_reports_card_admin",
        ["cards.manage"],
    )

    organization_service = OrganizationService(db_session)
    root = organization_service.create_root_for_actor(
        actor_user_id=system_admin.id,
        code="api-reports-root",
        name="API Reports Root",
    )
    child = organization_service.create_child(
        parent_id=root.id,
        code="api-reports-child",
        name="API Reports Child",
        created_by=system_admin.id,
    )
    sibling = organization_service.create_child(
        parent_id=root.id,
        code="api-reports-sibling",
        name="API Reports Sibling",
        created_by=system_admin.id,
    )

    schema_service = RegistrySchemaService(db_session)
    registry = schema_service.create_registry_for_actor(
        actor_user_id=system_admin.id,
        code="api-reports-registry",
        name="API Reports Registry",
    )
    block = schema_service.create_block_for_actor(
        actor_user_id=system_admin.id,
        registry_id=registry.id,
        code="main",
        title="Main",
    )
    title_field = schema_service.create_field_for_actor(
        actor_user_id=system_admin.id,
        block_id=block.id,
        code="title",
        label="Title",
        field_type="text",
    )

    _grant_access(
        db_session,
        user_id=schema_admin.id,
        role_id=schema_role.id,
        registry_id=registry.id,
        created_by=system_admin.id,
    )
    _grant_access(
        db_session,
        user_id=card_admin.id,
        role_id=card_role.id,
        organization_id=child.id,
        registry_id=registry.id,
        include_descendants=True,
        created_by=system_admin.id,
    )
    _grant_access(
        db_session,
        user_id=sibling_admin.id,
        role_id=card_role.id,
        organization_id=sibling.id,
        registry_id=registry.id,
        include_descendants=True,
        created_by=system_admin.id,
    )

    card_service = CardService(db_session)
    child_card = card_service.create_card_for_actor(
        actor_user_id=card_admin.id,
        registry_id=registry.id,
        organization_id=child.id,
        display_name="Visible report card",
    )
    sibling_card = card_service.create_card_for_actor(
        actor_user_id=sibling_admin.id,
        registry_id=registry.id,
        organization_id=sibling.id,
        display_name="Hidden sibling report card",
    )
    card_service.set_field_value_for_actor(
        actor_user_id=card_admin.id,
        card_id=child_card.id,
        field_id=title_field.id,
        value="Visible field value",
    )
    card_service.set_field_value_for_actor(
        actor_user_id=sibling_admin.id,
        card_id=sibling_card.id,
        field_id=title_field.id,
        value="Hidden field value",
    )

    return {
        "system_admin": system_admin,
        "schema_admin": schema_admin,
        "card_admin": card_admin,
        "sibling_admin": sibling_admin,
        "registry": registry,
        "child": child,
        "sibling": sibling,
        "child_card": child_card,
        "sibling_card": sibling_card,
    }


def _create_report_template(
    api_client: TestClient,
    *,
    registry_id: UUID,
    actor_user_id: UUID,
    code: str,
    report_type: str,
    output_format: str = "json",
) -> dict[str, Any]:
    response = api_client.post(
        f"/api/v1/registries/{registry_id}/report-templates",
        headers=_actor_headers(actor_user_id),
        json={
            "code": code,
            "name": f"{code} report",
            "report_type": report_type,
            "output_format": output_format,
        },
    )
    assert response.status_code == 201, response.text
    payload = response.json()
    assert payload["code"] == code
    assert payload["report_type"] == report_type
    assert payload["output_format"] == output_format
    assert "stored_file_id" not in payload
    assert "storage_key" not in payload
    return payload


def _extract_pdf_text(content: bytes) -> str:
    reader = PdfReader(BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def test_xlsx_report_output_renderer_creates_workbook_bytes() -> None:
    template = ReportTemplate(
        code="registry-cards-xlsx",
        name="Registry cards XLSX",
        report_type="registry_cards",
        output_format="xlsx",
    )
    rendered = _RenderedReport(
        report_type="registry_cards",
        card_id=None,
        content={
            "cards": [
                {
                    "id": "card-1",
                    "registry_id": "registry-1",
                    "organization_id": "organization-1",
                    "org_unit_id": None,
                    "display_name": "Visible report card",
                    "lifecycle_status": "draft",
                    "created_at": "2026-06-30T00:00:00+00:00",
                }
            ]
        },
        summary={"card_count": 1},
        row_count=1,
    )

    output = ReportService(session=None, storage=None)._render_report_output(  # type: ignore[arg-type]
        template=template,
        rendered=rendered,
    )

    assert (
        output.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert output.filename.endswith(".xlsx")
    workbook = load_workbook(BytesIO(output.content), read_only=True)
    worksheet = workbook["registry_cards"]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "id",
        "registry_id",
        "organization_id",
        "org_unit_id",
        "display_name",
        "lifecycle_status",
        "created_at",
    )
    assert rows[1][4] == "Visible report card"


def test_pdf_report_output_renderer_creates_pdf_bytes() -> None:
    template = ReportTemplate(
        code="registry-cards-pdf",
        name="Registry cards PDF",
        report_type="registry_cards",
        output_format="pdf",
    )
    rendered = _RenderedReport(
        report_type="registry_cards",
        card_id=None,
        content={
            "cards": [
                {
                    "id": "card-1",
                    "registry_id": "registry-1",
                    "organization_id": "organization-1",
                    "org_unit_id": None,
                    "display_name": "Visible report card",
                    "lifecycle_status": "draft",
                    "created_at": "2026-06-30T00:00:00+00:00",
                }
            ]
        },
        summary={"card_count": 1},
        row_count=1,
    )

    output = ReportService(session=None, storage=None)._render_report_output(  # type: ignore[arg-type]
        template=template,
        rendered=rendered,
    )

    assert output.content_type == "application/pdf"
    assert output.filename.endswith(".pdf")
    extracted_text = _extract_pdf_text(output.content)
    assert "Registry cards PDF" in extracted_text
    assert "Visible report card" in extracted_text
    assert "card_count" in extracted_text


def test_report_template_update_service_accepts_type_and_format(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class FakeSession:
        def __init__(self) -> None:
            self.flushed = False

        def flush(self) -> None:
            self.flushed = True

    class FakeAuditService:
        def __init__(self, session: FakeSession) -> None:
            self.session = session
            self.events: list[dict[str, Any]] = []

        def record_user_event(self, **kwargs: Any) -> None:
            self.events.append(kwargs)

    session = FakeSession()
    audit_service = FakeAuditService(session)
    template = ReportTemplate(
        id=uuid4(),
        registry_id=uuid4(),
        code="editable-report",
        name="Editable report",
        report_type="registry_cards",
        output_format="json",
    )
    service = ReportService(session=session, storage=None)  # type: ignore[arg-type]
    monkeypatch.setattr(service, "_get_active_template", lambda template_id: template)
    monkeypatch.setattr(
        service,
        "_require_schema_permission",
        lambda actor_user_id, registry_id: None,
    )
    monkeypatch.setattr("app.services.reports.AuditService", lambda session: audit_service)

    updated = service.update_template_for_actor(
        actor_user_id=uuid4(),
        template_id=template.id,
        updates={"report_type": "period_summary", "output_format": "pdf"},
    )

    assert updated.report_type == "period_summary"
    assert updated.output_format == "pdf"
    assert session.flushed is True
    assert audit_service.events[0]["old_data_json"] == {
        "report_type": "registry_cards",
        "output_format": "json",
    }
    assert audit_service.events[0]["new_data_json"] == {
        "report_type": "period_summary",
        "output_format": "pdf",
    }


def test_report_template_settings_can_be_updated_and_audited(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _report_api_context(db_session)
    template = _create_report_template(
        api_client,
        registry_id=context["registry"].id,
        actor_user_id=context["schema_admin"].id,
        code="editable-template",
        report_type="registry_cards",
    )

    update_response = api_client.patch(
        f"/api/v1/report-templates/{template['id']}",
        headers=_actor_headers(context["schema_admin"].id),
        json={
            "name": "Updated registry report",
            "description": "Updated report description",
            "report_type": "period_summary",
            "default_parameters_json": {"limit": 25},
            "output_format": "pdf",
        },
    )
    assert update_response.status_code == 200, update_response.text
    updated = update_response.json()
    assert updated["id"] == template["id"]
    assert updated["code"] == "editable-template"
    assert updated["report_type"] == "period_summary"
    assert updated["output_format"] == "pdf"
    assert updated["name"] == "Updated registry report"
    assert updated["description"] == "Updated report description"
    assert updated["default_parameters_json"] == {"limit": 25}

    list_response = api_client.get(
        f"/api/v1/registries/{context['registry'].id}/report-templates",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert list_response.status_code == 200, list_response.text
    assert list_response.json()["items"][0]["name"] == "Updated registry report"

    forbidden_response = api_client.patch(
        f"/api/v1/report-templates/{template['id']}",
        headers=_actor_headers(context["card_admin"].id),
        json={"name": "Forbidden update"},
    )
    assert forbidden_response.status_code == 403, forbidden_response.text

    archive_response = api_client.delete(
        f"/api/v1/report-templates/{template['id']}",
        headers=_actor_headers(context["schema_admin"].id),
    )
    assert archive_response.status_code == 200, archive_response.text
    archived_update_response = api_client.patch(
        f"/api/v1/report-templates/{template['id']}",
        headers=_actor_headers(context["schema_admin"].id),
        json={"name": "Archived update"},
    )
    assert archived_update_response.status_code == 400, archived_update_response.text

    invalid_format_template = _create_report_template(
        api_client,
        registry_id=context["registry"].id,
        actor_user_id=context["schema_admin"].id,
        code="invalid-format-template",
        report_type="registry_cards",
    )
    invalid_format_response = api_client.patch(
        f"/api/v1/report-templates/{invalid_format_template['id']}",
        headers=_actor_headers(context["schema_admin"].id),
        json={"output_format": "xml"},
    )
    assert invalid_format_response.status_code == 400, invalid_format_response.text
    assert "Unsupported report output format" in invalid_format_response.text

    audit_actions = set(
        db_session.scalars(
            select(AuditEvent.action).where(
                AuditEvent.object_type == "report_template",
                AuditEvent.object_id == UUID(template["id"]),
            )
        ).all()
    )
    assert {"report_template_create", "report_template_update", "report_template_archive"} <= (
        audit_actions
    )


def test_csv_registry_report_runs_are_scoped_stored_and_downloadable(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _report_api_context(db_session)
    csv_template = _create_report_template(
        api_client,
        registry_id=context["registry"].id,
        actor_user_id=context["schema_admin"].id,
        code="registry-cards-csv",
        report_type="registry_cards",
        output_format="csv",
    )

    run_response = api_client.post(
        f"/api/v1/report-templates/{csv_template['id']}/runs",
        headers=_actor_headers(context["card_admin"].id),
        json={"parameters": {"organization_id": str(context["child"].id)}},
    )
    assert run_response.status_code == 201, run_response.text
    run_payload = run_response.json()
    assert run_payload["report_template_id"] == csv_template["id"]
    assert run_payload["report_type"] == "registry_cards"
    assert run_payload["row_count"] == 1
    assert run_payload["output_content_type"] == "text/csv; charset=utf-8"
    assert run_payload["output_filename"].endswith(".csv")

    stored_file = db_session.scalar(
        select(StoredFile).where(StoredFile.original_filename == run_payload["output_filename"])
    )
    assert stored_file is not None
    assert stored_file.storage_key.startswith("reports/")
    assert stored_file.content_type == "text/csv; charset=utf-8"
    assert stored_file.scanner_details_json == {
        "source": "report_run_v1",
        "report_type": "registry_cards",
        "output_format": "csv",
    }

    download_response = api_client.get(
        f"/api/v1/report-runs/{run_payload['id']}/content",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert download_response.status_code == 200, download_response.text
    assert download_response.headers["content-type"].startswith("text/csv")
    assert "attachment;" in download_response.headers["content-disposition"]
    assert download_response.headers["x-report-filename"].endswith(".csv")
    csv_text = download_response.content.decode("utf-8")
    rows = list(csv.DictReader(StringIO(csv_text)))
    assert len(rows) == 1
    assert rows[0]["id"] == str(context["child_card"].id)
    assert rows[0]["display_name"] == "Visible report card"
    assert rows[0]["lifecycle_status"] == "draft"
    assert "Hidden sibling report card" not in csv_text

    audit_actions = set(
        db_session.scalars(
            select(AuditEvent.action).where(AuditEvent.object_type == "report_run")
        ).all()
    )
    assert {"report_run_generate", "report_run_download"} <= audit_actions


def test_xlsx_registry_report_runs_are_scoped_stored_and_downloadable(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _report_api_context(db_session)
    xlsx_template = _create_report_template(
        api_client,
        registry_id=context["registry"].id,
        actor_user_id=context["schema_admin"].id,
        code="registry-cards-xlsx",
        report_type="registry_cards",
        output_format="xlsx",
    )

    run_response = api_client.post(
        f"/api/v1/report-templates/{xlsx_template['id']}/runs",
        headers=_actor_headers(context["card_admin"].id),
        json={"parameters": {"organization_id": str(context["child"].id)}},
    )
    assert run_response.status_code == 201, run_response.text
    run_payload = run_response.json()
    assert run_payload["report_template_id"] == xlsx_template["id"]
    assert run_payload["report_type"] == "registry_cards"
    assert run_payload["row_count"] == 1
    assert (
        run_payload["output_content_type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert run_payload["output_filename"].endswith(".xlsx")

    stored_file = db_session.scalar(
        select(StoredFile).where(StoredFile.original_filename == run_payload["output_filename"])
    )
    assert stored_file is not None
    assert stored_file.storage_key.startswith("reports/")
    assert (
        stored_file.content_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert stored_file.scanner_details_json == {
        "source": "report_run_v1",
        "report_type": "registry_cards",
        "output_format": "xlsx",
    }

    download_response = api_client.get(
        f"/api/v1/report-runs/{run_payload['id']}/content",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert download_response.status_code == 200, download_response.text
    assert (
        download_response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in download_response.headers["content-disposition"]
    assert download_response.headers["x-report-filename"].endswith(".xlsx")

    workbook = load_workbook(BytesIO(download_response.content), read_only=True)
    worksheet = workbook["registry_cards"]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "id",
        "registry_id",
        "organization_id",
        "org_unit_id",
        "display_name",
        "lifecycle_status",
        "created_at",
    )
    assert rows[1][0] == str(context["child_card"].id)
    assert rows[1][4] == "Visible report card"
    assert rows[1][5] == "draft"
    flattened = "\n".join(str(cell) for row in rows for cell in row if cell is not None)
    assert "Hidden sibling report card" not in flattened

    audit_actions = set(
        db_session.scalars(
            select(AuditEvent.action).where(AuditEvent.object_type == "report_run")
        ).all()
    )
    assert {"report_run_generate", "report_run_download"} <= audit_actions


def test_pdf_registry_report_runs_are_scoped_stored_and_downloadable(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _report_api_context(db_session)
    pdf_template = _create_report_template(
        api_client,
        registry_id=context["registry"].id,
        actor_user_id=context["schema_admin"].id,
        code="registry-cards-pdf",
        report_type="registry_cards",
        output_format="pdf",
    )

    run_response = api_client.post(
        f"/api/v1/report-templates/{pdf_template['id']}/runs",
        headers=_actor_headers(context["card_admin"].id),
        json={"parameters": {"organization_id": str(context["child"].id)}},
    )
    assert run_response.status_code == 201, run_response.text
    run_payload = run_response.json()
    assert run_payload["report_template_id"] == pdf_template["id"]
    assert run_payload["report_type"] == "registry_cards"
    assert run_payload["row_count"] == 1
    assert run_payload["output_content_type"] == "application/pdf"
    assert run_payload["output_filename"].endswith(".pdf")

    stored_file = db_session.scalar(
        select(StoredFile).where(StoredFile.original_filename == run_payload["output_filename"])
    )
    assert stored_file is not None
    assert stored_file.storage_key.startswith("reports/")
    assert stored_file.content_type == "application/pdf"
    assert stored_file.scanner_details_json == {
        "source": "report_run_v1",
        "report_type": "registry_cards",
        "output_format": "pdf",
    }

    download_response = api_client.get(
        f"/api/v1/report-runs/{run_payload['id']}/content",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert download_response.status_code == 200, download_response.text
    assert download_response.headers["content-type"] == "application/pdf"
    assert "attachment;" in download_response.headers["content-disposition"]
    assert download_response.headers["x-report-filename"].endswith(".pdf")
    extracted_text = _extract_pdf_text(download_response.content)
    assert "registry-cards-pdf report" in extracted_text
    assert "Visible report card" in extracted_text
    assert "Hidden sibling report card" not in extracted_text

    audit_actions = set(
        db_session.scalars(
            select(AuditEvent.action).where(AuditEvent.object_type == "report_run")
        ).all()
    )
    assert {"report_run_generate", "report_run_download"} <= audit_actions


def test_report_runs_list_newest_runs_first(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _report_api_context(db_session)
    template = _create_report_template(
        api_client,
        registry_id=context["registry"].id,
        actor_user_id=context["schema_admin"].id,
        code="run-order",
        report_type="registry_cards",
    )

    first_response = api_client.post(
        f"/api/v1/report-templates/{template['id']}/runs",
        headers=_actor_headers(context["card_admin"].id),
        json={"parameters": {"organization_id": str(context["child"].id), "q": "first"}},
    )
    assert first_response.status_code == 201, first_response.text
    second_response = api_client.post(
        f"/api/v1/report-templates/{template['id']}/runs",
        headers=_actor_headers(context["card_admin"].id),
        json={"parameters": {"organization_id": str(context["child"].id), "q": "second"}},
    )
    assert second_response.status_code == 201, second_response.text
    first_payload = first_response.json()
    second_payload = second_response.json()

    first_run = db_session.get(ReportRun, UUID(first_payload["id"]))
    second_run = db_session.get(ReportRun, UUID(second_payload["id"]))
    assert first_run is not None
    assert second_run is not None
    first_run.created_at = datetime(2026, 1, 1, tzinfo=UTC)
    second_run.created_at = datetime(2026, 1, 2, tzinfo=UTC)
    db_session.flush()

    list_response = api_client.get(
        f"/api/v1/registries/{context['registry'].id}/report-runs",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert list_response.status_code == 200, list_response.text
    assert [item["id"] for item in list_response.json()["items"]] == [
        second_payload["id"],
        first_payload["id"],
    ]


def test_registry_and_period_report_runs_are_scoped_stored_and_audited(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _report_api_context(db_session)
    registry_template = _create_report_template(
        api_client,
        registry_id=context["registry"].id,
        actor_user_id=context["schema_admin"].id,
        code="registry-cards",
        report_type="registry_cards",
    )

    list_response = api_client.get(
        f"/api/v1/registries/{context['registry'].id}/report-templates",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert list_response.status_code == 200, list_response.text
    assert [item["id"] for item in list_response.json()["items"]] == [registry_template["id"]]

    run_response = api_client.post(
        f"/api/v1/report-templates/{registry_template['id']}/runs",
        headers=_actor_headers(context["card_admin"].id),
        json={"parameters": {"organization_id": str(context["child"].id)}},
    )
    assert run_response.status_code == 201, run_response.text
    run_payload = run_response.json()
    assert run_payload["report_template_id"] == registry_template["id"]
    assert run_payload["registry_id"] == str(context["registry"].id)
    assert run_payload["report_type"] == "registry_cards"
    assert run_payload["run_status"] == "generated"
    assert run_payload["row_count"] == 1
    assert run_payload["output_content_type"] == "application/json"
    assert run_payload["output_filename"].endswith(".json")
    assert run_payload["summary_json"] == {"card_count": 1}
    assert "stored_file_id" not in run_payload
    assert "checksum_sha256" not in run_payload
    assert "storage_key" not in run_payload

    stored_file = db_session.scalar(
        select(StoredFile).where(StoredFile.original_filename == run_payload["output_filename"])
    )
    assert stored_file is not None
    assert stored_file.storage_key.startswith("reports/")
    assert stored_file.scanner_details_json == {
        "source": "report_run_v1",
        "report_type": "registry_cards",
        "output_format": "json",
    }

    download_response = api_client.get(
        f"/api/v1/report-runs/{run_payload['id']}/content",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert download_response.status_code == 200, download_response.text
    assert download_response.headers["content-type"] == "application/json"
    assert "attachment;" in download_response.headers["content-disposition"]
    report_content = json.loads(download_response.content)
    assert report_content["format_version"] == "report_run_v1"
    assert report_content["report_type"] == "registry_cards"
    assert [card["id"] for card in report_content["cards"]] == [str(context["child_card"].id)]
    assert "Hidden sibling report card" not in download_response.text

    runs_list_response = api_client.get(
        f"/api/v1/registries/{context['registry'].id}/report-runs",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert runs_list_response.status_code == 200, runs_list_response.text
    assert [item["id"] for item in runs_list_response.json()["items"]] == [run_payload["id"]]

    period_template = _create_report_template(
        api_client,
        registry_id=context["registry"].id,
        actor_user_id=context["schema_admin"].id,
        code="period-summary",
        report_type="period_summary",
    )
    period_response = api_client.post(
        f"/api/v1/report-templates/{period_template['id']}/runs",
        headers=_actor_headers(context["card_admin"].id),
        json={
            "parameters": {
                "created_from": "2000-01-01T00:00:00+00:00",
                "created_to": "2999-01-01T00:00:00+00:00",
            }
        },
    )
    assert period_response.status_code == 201, period_response.text
    period_payload = period_response.json()
    assert period_payload["report_type"] == "period_summary"
    assert period_payload["summary_json"] == {
        "card_count": 1,
        "lifecycle_status_counts": {"draft": 1},
    }

    audit_actions = set(
        db_session.scalars(
            select(AuditEvent.action).where(
                AuditEvent.object_type.in_(("report_template", "report_run"))
            )
        ).all()
    )
    assert {"report_template_create", "report_run_generate", "report_run_download"} <= audit_actions


def test_card_detail_report_respects_card_scope_and_archive(
    api_client: TestClient,
    db_session: Session,
) -> None:
    context = _report_api_context(db_session)
    card_template = _create_report_template(
        api_client,
        registry_id=context["registry"].id,
        actor_user_id=context["schema_admin"].id,
        code="card-detail",
        report_type="card_detail",
    )

    run_response = api_client.post(
        f"/api/v1/report-templates/{card_template['id']}/runs",
        headers=_actor_headers(context["card_admin"].id),
        json={"parameters": {"card_id": str(context["child_card"].id)}},
    )
    assert run_response.status_code == 201, run_response.text
    run_payload = run_response.json()
    assert run_payload["card_id"] == str(context["child_card"].id)
    assert run_payload["row_count"] == 1

    download_response = api_client.get(
        f"/api/v1/report-runs/{run_payload['id']}/content",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert download_response.status_code == 200, download_response.text
    report_content = json.loads(download_response.content)
    assert report_content["card"]["id"] == str(context["child_card"].id)
    assert (
        report_content["card"]["blocks"]["main"]["instances"][0]["fields"]["title"]["value"]
        == "Visible field value"
    )

    forbidden_response = api_client.post(
        f"/api/v1/report-templates/{card_template['id']}/runs",
        headers=_actor_headers(context["card_admin"].id),
        json={"parameters": {"card_id": str(context["sibling_card"].id)}},
    )
    assert forbidden_response.status_code == 403, forbidden_response.text

    archive_response = api_client.delete(
        f"/api/v1/report-runs/{run_payload['id']}",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert archive_response.status_code == 200, archive_response.text
    assert archive_response.json()["archived_at"] is not None

    active_runs = api_client.get(
        f"/api/v1/registries/{context['registry'].id}/report-runs",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert active_runs.status_code == 200, active_runs.text
    assert active_runs.json()["items"] == []

    archived_runs = api_client.get(
        f"/api/v1/registries/{context['registry'].id}/report-runs?include_archive=true",
        headers=_actor_headers(context["card_admin"].id),
    )
    assert archived_runs.status_code == 200, archived_runs.text
    assert [item["id"] for item in archived_runs.json()["items"]] == [run_payload["id"]]

    audit_actions = set(
        db_session.scalars(
            select(AuditEvent.action).where(AuditEvent.object_type == "report_run")
        ).all()
    )
    assert {"report_run_generate", "report_run_download", "report_run_archive"} <= audit_actions
